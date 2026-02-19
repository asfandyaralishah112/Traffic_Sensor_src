import tkinter as tk
from tkinter import ttk, messagebox
import serial
import serial.tools.list_ports
import json
import os
import random
import subprocess
import threading
import time
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from datetime import datetime
import sys

CONFIG_FILE = "provisioning_config.json"
EXCEL_FILE = "device_registry.xlsx"

FIRMWARE_DIR = r"src\build\esp32.esp32.esp32c6"

BOOTLOADER = os.path.join(FIRMWARE_DIR, "src.ino.bootloader.bin")
PARTITIONS = os.path.join(FIRMWARE_DIR, "src.ino.partitions.bin")
FIRMWARE   = os.path.join(FIRMWARE_DIR, "src.ino.bin")

MAX_RETRIES = 5


class ProvisionTool:

    def __init__(self, root):
        self.root = root
        self.root.title("Cavline Global - Factory Tool")
        self.root.geometry("580x700")
        self.root.configure(bg="#071218")

        self.create_ui()
        self.refresh_ports()
        self.load_config()

    # ==================================================
    # UI
    # ==================================================
    def create_ui(self):

        container = tk.Frame(self.root, bg="#071218")
        container.pack(fill="both", expand=True, padx=20, pady=15)

        try:
            img = Image.open("logo.png")
            img = img.resize((260, int(260 * img.height / img.width)))
            self.logo_img = ImageTk.PhotoImage(img)
            tk.Label(container, image=self.logo_img,
                     bg="#071218").pack(pady=(0, 10))
        except:
            pass

        tk.Label(container,
                 text="Traffic Sensor Factory Provisioning",
                 bg="#071218",
                 fg="#e6eef3",
                 font=("Segoe UI", 14, "bold")).pack(pady=(0, 10))

        form = tk.Frame(container, bg="#071218")
        form.pack(fill="x")

        self.entries = {}

        fields = [
            ("Start UID", tk.Entry),
            ("COM Port", ttk.Combobox),
            ("MQTT Server", tk.Entry),
            ("MQTT Port", tk.Entry),
        ]

        for i, (label, widget) in enumerate(fields):
            tk.Label(form, text=label,
                     bg="#071218", fg="#e6eef3").grid(
                row=i, column=0, sticky="w", pady=6)

            if widget == ttk.Combobox:
                entry = widget(form, width=30, state="readonly")
            else:
                entry = widget(form, width=32)

            entry.grid(row=i, column=1, pady=6, padx=(10, 0))
            self.entries[label] = entry

        ttk.Button(form,
                   text="Refresh Ports",
                   command=self.refresh_ports).grid(row=1, column=2, padx=10)

        self.upload_btn = tk.Button(container,
                                    text="FLASH + PROVISION",
                                    command=self.start_thread,
                                    bg="#1c7ed6",
                                    fg="white",
                                    font=("Segoe UI", 11, "bold"),
                                    relief="flat",
                                    height=2)
        self.upload_btn.pack(fill="x", pady=10)

        tk.Label(container, text="Output",
                 bg="#071218", fg="#8fa3ad").pack(anchor="w")

        self.output = tk.Text(container,
                              height=15,
                              bg="#0f171d",
                              fg="#e6eef3",
                              insertbackground="white")
        self.output.pack(fill="both", expand=True)

        self.progress = ttk.Progressbar(container,
                                        orient="horizontal",
                                        mode="determinate")
        self.progress.pack(fill="x", pady=10)

    # ==================================================
    # Logging & Progress
    # ==================================================
    def log(self, text):
        self.output.insert(tk.END, text + "\n")
        self.output.see(tk.END)
        self.root.update_idletasks()

    def set_progress(self, value):
        self.progress["value"] = value
        self.root.update_idletasks()

    # ==================================================
    # COM Ports
    # ==================================================
    def refresh_ports(self):
        ports = serial.tools.list_ports.comports()
        port_list = [p.device for p in ports]
        combo = self.entries["COM Port"]
        combo["values"] = port_list
        if port_list:
            combo.current(0)

    # ==================================================
    # Flash Firmware
    # ==================================================
    def flash_firmware(self, port):

        self.log("Flashing ESP32-C6 firmware...")
        self.set_progress(10)

        command = [
            sys.executable,
            "-m",
            "esptool",
            "--chip", "esp32c6",
            "--port", port,
            "--baud", "921600",
            "--before", "default_reset",
            "--after", "hard_reset",
            "write_flash",
            "0x0", BOOTLOADER,
            "0x8000", PARTITIONS,
            "0x10000", FIRMWARE
        ]

        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True
        )

        for line in process.stdout:
            self.log(line.strip())

        process.wait()

        if process.returncode == 0:
            self.log("Flash successful.")
            self.set_progress(50)
            return True
        else:
            self.log("Flash failed.")
            return False

    # ==================================================
    # Retry Provisioning
    # ==================================================
    def write_config_with_retry(self, port, payload):

        for attempt in range(1, MAX_RETRIES + 1):

            self.log(f"Provision attempt {attempt}/{MAX_RETRIES}...")
            self.set_progress(50 + attempt * 5)

            try:
                ser = serial.Serial(port, 115200, timeout=3)
                time.sleep(1)
                ser.reset_input_buffer()

                ser.write((json.dumps(payload) + "\n").encode())

                response = ser.readline().decode().strip()
                ser.close()

                if response == "OK":
                    self.log("ESP confirmed configuration.")
                    return True
                else:
                    self.log("No confirmation. Retrying...")
                    time.sleep(1)

            except Exception as e:
                self.log(f"Attempt failed: {str(e)}")
                time.sleep(1)

        return False

    # ==================================================
    # Provision Flow
    # ==================================================
    def start_thread(self):
        threading.Thread(target=self.provision).start()

    def provision(self):

        self.progress["value"] = 0
        self.output.delete("1.0", tk.END)

        uid = self.entries["Start UID"].get()
        port = self.entries["COM Port"].get()

        if not uid or not port:
            messagebox.showerror("Error", "UID and COM Port required")
            return

        if not self.flash_firmware(port):
            messagebox.showerror("Error", "Firmware flashing failed")
            return

        self.log("Waiting for ESP reboot...")
        time.sleep(2)

        mqtt_user = uid
        mqtt_pass = self.generate_password()

        payload = {
            "cmd": "write_config",
            "uid": uid,
            "mqtt_server": self.entries["MQTT Server"].get(),
            "mqtt_port": int(self.entries["MQTT Port"].get()),
            "mqtt_user": mqtt_user,
            "mqtt_pass": mqtt_pass
        }

        success = self.write_config_with_retry(port, payload)

        if not success:
            messagebox.showerror("Error",
                                 "Provisioning failed after 5 retries")
            return

        self.log("Logging to Excel...")
        self.set_progress(95)

        self.log_to_excel(uid, mqtt_user,
                          mqtt_pass,
                          self.entries["MQTT Server"].get())

        new_uid = self.increment_uid(uid)
        self.entries["Start UID"].delete(0, tk.END)
        self.entries["Start UID"].insert(0, new_uid)

        self.save_config()

        self.set_progress(100)
        self.log("Provisioning completed successfully.")
        messagebox.showinfo("Success", f"Provisioned:\n{uid}")

    # ==================================================
    # Utilities
    # ==================================================
    def increment_uid(self, uid):
        prefix, number = uid.rsplit("-", 1)
        new_number = str(int(number) + 1).zfill(6)
        return f"{prefix}-{new_number}"

    def generate_password(self, length=12):
        chars = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
        return ''.join(random.choice(chars) for _ in range(length))

    def log_to_excel(self, uid, user, password, server):
        now = datetime.now()

        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["UID", "MQTT User",
                       "MQTT Password", "MQTT Server",
                       "Date", "Time"])
            wb.save(EXCEL_FILE)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        ws.append([
            uid,
            user,
            password,
            server,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S")
        ])

        wb.save(EXCEL_FILE)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as f:
                data = json.load(f)
                for k in self.entries:
                    if k in data:
                        self.entries[k].insert(0, data[k])

    def save_config(self):
        data = {k: self.entries[k].get() for k in self.entries}
        with open(CONFIG_FILE, "w") as f:
            json.dump(data, f, indent=4)


# ==================================================
# START
# ==================================================
root = tk.Tk()
app = ProvisionTool(root)
root.mainloop()
